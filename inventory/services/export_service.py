import io
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class ExportService:
    """
    导出服务类，用于将数据导出为Excel文件
    """
    
    @staticmethod
    def export_to_excel(data, filename, sheet_name='Sheet1'):
        """
        将数据导出为Excel文件
        :param data: 数据，格式为[{'header1': value1, 'header2': value2, ...}, ...]
        :param filename: 文件名
        :param sheet_name: 工作表名称
        :return: HttpResponse对象
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # 设置列标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置单元格边框
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 写入表头
        if data and len(data) > 0:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                    cell.border = thin_border
                    # 日期格式处理
                    if isinstance(cell.value, datetime.datetime):
                        cell.number_format = 'yyyy-mm-dd'
                    elif isinstance(cell.value, datetime.date):
                        cell.number_format = 'yyyy-mm-dd'
            
            # 自动调整列宽
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                # 设置一个基本宽度，然后根据内容增加
                column_width = max(10, len(str(header)) + 2)  # 表头宽度 + 2
                
                # 检查列内容，找到最长的内容
                for row_idx in range(2, len(data) + 2):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        column_width = max(column_width, min(len(str(cell_value)) + 2, 50))  # 最大宽度50
                
                worksheet.column_dimensions[column_letter].width = column_width
        
        # 创建响应
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def format_member_data_for_export(member_data, start_date, end_date):
        """
        格式化会员分析数据用于导出
        """
        # 头部信息
        header_data = [
            {'名称': '报表类型', '数值': '会员分析报表'},
            {'名称': '统计周期', '数值': f"{start_date} 至 {end_date}"},
            {'名称': '会员总数', '数值': member_data['total_members']},
            {'名称': '新增会员数', '数值': member_data['new_members']},
            {'名称': '活跃会员数', '数值': member_data['active_members']},
            {'名称': '会员活跃率', '数值': f"{member_data['activity_rate']}%"},
            {'名称': '', '数值': ''},
        ]
        
        # 会员等级分布
        level_data = []
        for level in member_data['level_distribution']:
            level_data.append({
                '会员等级': level['level__name'],
                '会员数': level['count'],
                '占比': f"{(level['count'] / member_data['total_members'] * 100):.1f}%"
            })
        
        # 会员消费排行
        top_members_data = []
        for idx, member in enumerate(member_data['top_members'], 1):
            avg_order = member['period_spend'] / member['period_purchase_count'] if member['period_purchase_count'] > 0 else 0
            top_members_data.append({
                '排名': idx,
                '会员姓名': member['name'],
                '会员等级': member['level'].name,
                '手机号': member['phone'],
                '消费金额': f"¥{member['period_spend']:.2f}",
                '消费次数': member['period_purchase_count'],
                '平均客单价': f"¥{avg_order:.2f}"
            })
        
        return {
            'header': header_data,
            'level_distribution': level_data,
            'top_members': top_members_data
        }
    
    @staticmethod
    def export_member_analysis(member_data, start_date, end_date):
        """
        导出会员分析报表
        """
        formatted_data = ExportService.format_member_data_for_export(member_data, start_date, end_date)
        
        # 创建工作簿
        workbook = openpyxl.Workbook()
        
        # 头部信息表
        header_sheet = workbook.active
        header_sheet.title = "概览"
        header_sheet['A1'] = "会员分析报表"
        header_sheet['A1'].font = Font(bold=True, size=16)
        header_sheet.merge_cells('A1:B1')
        
        for idx, item in enumerate(formatted_data['header'], 3):
            header_sheet[f'A{idx}'] = item['名称']
            header_sheet[f'B{idx}'] = item['数值']
        
        # 会员等级分布表
        if formatted_data['level_distribution']:
            level_sheet = workbook.create_sheet(title="会员等级分布")
            # 表头
            level_sheet['A1'] = "会员等级"
            level_sheet['B1'] = "会员数"
            level_sheet['C1'] = "占比"
            
            # 表头样式
            for cell in level_sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            # 数据
            for idx, item in enumerate(formatted_data['level_distribution'], 2):
                level_sheet[f'A{idx}'] = item['会员等级']
                level_sheet[f'B{idx}'] = item['会员数']
                level_sheet[f'C{idx}'] = item['占比']
        
        # 会员消费排行表
        if formatted_data['top_members']:
            member_sheet = workbook.create_sheet(title="会员消费排行")
            # 表头
            headers = ["排名", "会员姓名", "会员等级", "手机号", "消费金额", "消费次数", "平均客单价"]
            for col, header in enumerate(headers, 1):
                member_sheet.cell(row=1, column=col, value=header)
                member_sheet.cell(row=1, column=col).font = Font(bold=True)
                member_sheet.cell(row=1, column=col).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            # 数据
            for row, item in enumerate(formatted_data['top_members'], 2):
                member_sheet.cell(row=row, column=1, value=item['排名'])
                member_sheet.cell(row=row, column=2, value=item['会员姓名'])
                member_sheet.cell(row=row, column=3, value=item['会员等级'])
                member_sheet.cell(row=row, column=4, value=item['手机号'])
                member_sheet.cell(row=row, column=5, value=item['消费金额'])
                member_sheet.cell(row=row, column=6, value=item['消费次数'])
                member_sheet.cell(row=row, column=7, value=item['平均客单价'])
        
        # 为每个工作表自动调整列宽
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 创建响应
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        # 格式化日期为字符串，用于文件名
        date_str = datetime.datetime.now().strftime('%Y%m%d')
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="会员分析报表_{date_str}.xlsx"'
        
        return response 